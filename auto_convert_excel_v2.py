import pandas as pd
from openpyxl import load_workbook

def convert_excel_file(input_path: str, output_path: str):
    # 1. Load file
    df = pd.read_excel(input_path)

    # 2. Column mapping
    column_mapping = {
        "광고 이름": "제목",
        "지출 금액 (KRW)": "광고비",
        "구매": "구매",
        "구매 전환값": "매출",
        "구매 ROAS(광고 지출 대비 수익률)": "ROAS",
        "CPC(전체) (KRW)": "CPC",
        "전환율(CVR)": "CVR",
        "CTR(전체)": "CTR",
        "클릭(전체)": "클릭",
        "동영상 재생": "동영상 재생",
        "동영상 3초 이상 재생": "동영상 3초 이상 재생",
        "동영상 100% 재생": "동영상 100% 재생",
    }
    df = df[list(column_mapping.keys())].rename(columns=column_mapping)

    # 3. 광고비 0 제거
    df = df[df["광고비"] > 0].copy()

    # 4. 후크 및 지속 계산 (0.01 보정 적용)
    df["후크"] = (df["동영상 3초 이상 재생"] / df["동영상 재생"]).round(4) * 0.01
    df["지속"] = (df["동영상 100% 재생"] / df["동영상 3초 이상 재생"]).round(4) * 0.01

    # 5. 컬럼 순서 정리
    columns = ["제목", "광고비", "구매", "매출", "ROAS", "CPC", "CVR", "CTR", "클릭",
               "후크", "지속", "동영상 재생", "동영상 3초 이상 재생", "동영상 100% 재생"]
    df = df[columns]

    # 6. 엑셀로 저장 (임시)
    df.to_excel(output_path, index=False)

    # 7. 서식 적용
    wb = load_workbook(output_path)
    ws = wb.active
    col_idx = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        r = row[0].row

        # ROAS: 0.00
        ws.cell(r, col_idx["ROAS"]).number_format = "0.00"
        # CPC: 정수
        ws.cell(r, col_idx["CPC"]).number_format = "0"

        # CVR: 보정 + 퍼센트
        try:
            cvr_val = float(ws.cell(r, col_idx["CVR"]).value)
            if cvr_val >= 100:
                cvr_val *= 0.01
            ws.cell(r, col_idx["CVR"]).value = round(cvr_val, 4)
            ws.cell(r, col_idx["CVR"]).number_format = "0.00%"
        except:
            pass

        # CTR: 무조건 0.01 보정 후 퍼센트
        try:
            ctr_val = float(ws.cell(r, col_idx["CTR"]).value) * 0.01
            ws.cell(r, col_idx["CTR"]).value = round(ctr_val, 4)
            ws.cell(r, col_idx["CTR"]).number_format = "0.00%"
        except:
            pass

        # 후크 / 지속: 정수 % (이미 0.01 곱해졌으므로 100 곱하지 않음)
        for key in ["후크", "지속"]:
            try:
                val = float(ws.cell(r, col_idx[key]).value)
                ws.cell(r, col_idx[key]).value = round(val, 4)
                ws.cell(r, col_idx[key]).number_format = "0.00%"
            except:
                pass

    wb.save(output_path)
